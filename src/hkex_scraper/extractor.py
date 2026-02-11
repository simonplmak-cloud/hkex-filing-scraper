"""Document text and table extraction for PDF, HTML, and Excel files.

Outputs clean, structured Markdown suitable for LLM consumption and human reading.
Uses pymupdf4llm for PDF text extraction (headings, bold, lists) and camelot-py
for accurate table extraction (handles merged cells in HKEx regulatory forms).
Applies post-processing to remove page headers/footers, normalize bullets, and
fix spacing.
"""

from __future__ import annotations

import io
import os
import re
import tempfile
from typing import List, Tuple

from .utils import log, squash_ws

# ---------------------------------------------------------------------------
# Optional dependencies (graceful degradation)
# ---------------------------------------------------------------------------
try:
    import pymupdf  # type: ignore  # pip install PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    try:
        import fitz as pymupdf  # type: ignore  # older PyMuPDF
        PYMUPDF_AVAILABLE = True
    except ImportError:
        PYMUPDF_AVAILABLE = False

try:
    import pymupdf4llm  # type: ignore  # pip install pymupdf4llm
    PYMUPDF4LLM_AVAILABLE = True
except ImportError:
    PYMUPDF4LLM_AVAILABLE = False

try:
    import camelot  # type: ignore  # pip install camelot-py[cv]
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False

try:
    from bs4 import BeautifulSoup, Tag  # type: ignore  # pip install beautifulsoup4
    BS4_AVAILABLE = True
except ImportError:
    BS4_AVAILABLE = False

try:
    import openpyxl  # type: ignore  # pip install openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def check_dependencies() -> None:
    """Print warnings for missing optional extraction libraries."""
    if not PYMUPDF_AVAILABLE:
        print("WARNING: PyMuPDF not installed. PDF text extraction disabled. "
              "Run: pip install PyMuPDF")
    if not PYMUPDF4LLM_AVAILABLE:
        print("WARNING: pymupdf4llm not installed. PDF Markdown extraction degraded. "
              "Run: pip install pymupdf4llm")
    if not CAMELOT_AVAILABLE:
        print("WARNING: camelot-py not installed. PDF table extraction degraded "
              "(merged cells may be duplicated). Run: pip install camelot-py[cv]")
    if not BS4_AVAILABLE:
        print("WARNING: beautifulsoup4 not installed. HTML text extraction disabled. "
              "Run: pip install beautifulsoup4")
    if not OPENPYXL_AVAILABLE:
        print("WARNING: openpyxl not installed. Excel text extraction disabled. "
              "Run: pip install openpyxl")


# ---------------------------------------------------------------------------
# Markdown post-processing (shared across all extractors)
# ---------------------------------------------------------------------------

# Common HKEx PDF form codes that appear as page headers
_FORM_CODES = re.compile(
    r"^(?:FF\d{3}|EF\d{3}|MF\d{3}|SF\d{3}|CF\d{3})\s*$",
    re.MULTILINE,
)

# Page footer patterns: "Page X of Y", "v X.Y.Z", "Page X of Y v X.Y.Z"
_PAGE_FOOTERS = re.compile(
    r"^(?:Page\s+\d+\s+of\s+\d+(?:\s+v\s+[\d.]+)?|v\s+[\d.]+)\s*$",
    re.MULTILINE | re.IGNORECASE,
)

# Bullet normalization: various bullet characters -> standard "- "
_BULLET_CHARS = re.compile(r"^[\s]*[•●○▪▸►◆◇→⇒➤]\s*", re.MULTILINE)

# Numbered list normalization: "(1)" or "(a)" style -> "1." or "a."
_PAREN_NUMBERS = re.compile(r"^\((\d+)\)\s*", re.MULTILINE)
_PAREN_LETTERS = re.compile(r"^\(([a-z])\)\s*", re.MULTILINE)


def _clean_markdown(text: str) -> str:
    """Apply post-processing to clean up extracted Markdown text.

    Removes page headers/footers, normalizes bullets, collapses excessive
    blank lines, and trims trailing whitespace.
    """
    if not text:
        return ""

    # Remove HKEx form codes (e.g., "FF301")
    text = _FORM_CODES.sub("", text)

    # Remove page footers
    text = _PAGE_FOOTERS.sub("", text)

    # Normalize bullet characters to standard Markdown "- "
    text = _BULLET_CHARS.sub("- ", text)

    # Normalize parenthesized numbers/letters to standard list format
    text = _PAREN_NUMBERS.sub(r"\1. ", text)
    text = _PAREN_LETTERS.sub(r"\1. ", text)

    # Remove lines that are only dashes or underscores (decorative separators)
    text = re.sub(r"^[\s]*[-_=]{5,}[\s]*$", "", text, flags=re.MULTILINE)

    # Collapse 3+ consecutive blank lines to 2
    text = re.sub(r"\n{4,}", "\n\n\n", text)

    # Strip trailing whitespace from each line
    text = "\n".join(line.rstrip() for line in text.split("\n"))

    # Trim leading/trailing whitespace
    return text.strip()


# ---------------------------------------------------------------------------
# Markdown table helper (for HTML and Excel extractors)
# ---------------------------------------------------------------------------

def _table_to_markdown(headers: list, rows: list) -> str:
    """Convert a list of headers and rows into a Markdown table string."""
    if not headers and not rows:
        return ""
    if not headers or all(h is None for h in headers):
        ncols = len(rows[0]) if rows else 0
        headers = [f"Col{i + 1}" for i in range(ncols)]
    headers = [squash_ws(str(h)) if h else "" for h in headers]
    md_lines = ["| " + " | ".join(headers) + " |"]
    md_lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
    for row in rows:
        cells = [squash_ws(str(c)) if c else "" for c in row]
        while len(cells) < len(headers):
            cells.append("")
        cells = cells[: len(headers)]
        md_lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(md_lines)


def _extract_tables_from_md(md_text: str) -> List[dict]:
    """Parse Markdown tables from text and return structured table metadata.

    Scans the Markdown text for table blocks (lines starting with |) and
    extracts headers, row counts, and the raw Markdown for each table.
    """
    tables: List[dict] = []
    lines = md_text.split("\n")
    i = 0
    table_index = 0

    while i < len(lines):
        # Detect start of a Markdown table (line starts with |)
        if lines[i].strip().startswith("|"):
            table_lines: List[str] = []
            while i < len(lines) and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1

            # Need at least header + separator + 1 data row
            if len(table_lines) >= 3:
                table_index += 1
                header_line = table_lines[0]
                # Parse headers from first row
                headers = [
                    cell.strip()
                    for cell in header_line.strip("|").split("|")
                ]
                # Count data rows (skip header and separator)
                data_rows = [
                    ln for ln in table_lines[2:]
                    if ln.strip() and not re.match(r"^\|[\s\-:|]+\|$", ln.strip())
                ]
                tables.append({
                    "tableIndex": table_index,
                    "headers": headers,
                    "rowCount": len(data_rows),
                    "markdown": "\n".join(table_lines),
                })
        else:
            i += 1

    return tables


# ---------------------------------------------------------------------------
# Camelot-based PDF table extraction (handles merged cells correctly)
# ---------------------------------------------------------------------------

def _extract_tables_with_camelot(raw_bytes: bytes) -> List[dict]:
    """Extract tables from PDF bytes using camelot-py lattice mode.

    Camelot's lattice mode detects tables by looking for cell borders/lines,
    which is ideal for HKEx regulatory forms (FF301, FF305, etc.) that use
    bordered tables with merged cells.

    Returns a list of table dicts with headers, rows, accuracy, and Markdown.
    Falls back to an empty list on any error.
    """
    if not CAMELOT_AVAILABLE:
        return []

    # Camelot requires a file path, so write to a temp file
    fd, tmp_path = tempfile.mkstemp(suffix=".pdf")
    try:
        os.write(fd, raw_bytes)
        os.close(fd)

        tables = camelot.read_pdf(tmp_path, pages="all", flavor="lattice")
        log(f"    Camelot extracted {len(tables)} tables (lattice mode)")

        result: List[dict] = []
        for idx, tbl in enumerate(tables, start=1):
            df = tbl.df
            if df.empty or df.shape[0] < 2:
                continue

            # First row is typically the header
            raw_headers = [squash_ws(str(c)) for c in list(df.iloc[0])]

            # Clean up empty headers — replace blanks with positional names
            headers = []
            for i, h in enumerate(raw_headers):
                if h.strip():
                    headers.append(h)
                else:
                    headers.append(f"Col{i + 1}")

            # Data rows (skip header row)
            data_rows: list = []
            for row_idx in range(1, len(df)):
                row = [squash_ws(str(c)) if str(c).strip() else "" for c in list(df.iloc[row_idx])]
                # Skip rows that are completely empty
                if any(cell.strip() for cell in row):
                    data_rows.append(row)

            if not data_rows:
                continue

            # Build Markdown table
            md = _table_to_markdown(headers, data_rows)

            result.append({
                "tableIndex": idx,
                "page": tbl.page,
                "headers": headers,
                "rowCount": len(data_rows),
                "accuracy": round(tbl.accuracy, 1),
                "markdown": md,
            })

        return result

    except Exception as e:
        log(f"    Camelot table extraction error: {e}")
        return []
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass


# ---------------------------------------------------------------------------
# Table substitution: replace pymupdf4llm inline tables with camelot tables
# ---------------------------------------------------------------------------

def _strip_md_tables(md_text: str) -> Tuple[str, List[int]]:
    """Remove all Markdown tables from text, returning cleaned text and
    the character positions where tables were removed (for reinsertion).

    A Markdown table is a consecutive block of lines starting with '|',
    with at least 3 lines (header + separator + data).
    """
    lines = md_text.split("\n")
    result_lines: List[str] = []
    positions: List[int] = []  # char offset in result where each table was
    i = 0
    char_offset = 0

    while i < len(lines):
        if lines[i].strip().startswith("|"):
            # Collect the full table block
            table_start = i
            while i < len(lines) and lines[i].strip().startswith("|"):
                i += 1
            table_len = i - table_start
            if table_len >= 3:
                # This is a real table — record position and skip it
                positions.append(char_offset)
                # Add a placeholder blank line
                result_lines.append("")
                char_offset += 1  # for the newline
            else:
                # Too short to be a table — keep the lines
                for j in range(table_start, i):
                    result_lines.append(lines[j])
                    char_offset += len(lines[j]) + 1
        else:
            result_lines.append(lines[i])
            char_offset += len(lines[i]) + 1
            i += 1

    return "\n".join(result_lines), positions


def _substitute_tables(md_text: str, camelot_tables: List[dict]) -> str:
    """Replace pymupdf4llm inline tables in the Markdown text with cleaner
    camelot tables.

    Strategy:
    1. Strip all existing Markdown tables from the text
    2. Insert camelot tables at the positions where old tables were removed
    3. If there are more camelot tables than positions, append extras at the end
    4. If there are fewer camelot tables, some positions remain as blank lines
    """
    if not camelot_tables:
        return md_text

    stripped_text, positions = _strip_md_tables(md_text)

    # Build camelot table Markdown blocks
    camelot_mds = [t["markdown"] for t in camelot_tables if t.get("markdown")]

    if not camelot_mds:
        return md_text  # No camelot markdown to substitute

    # Split the stripped text at placeholder positions and interleave camelot tables
    lines = stripped_text.split("\n")
    result_lines: List[str] = []
    camelot_idx = 0
    i = 0
    pos_set = set()

    # Map positions back to line numbers
    char_offset = 0
    for line_num, line in enumerate(lines):
        if char_offset in positions:
            pos_set.add(line_num)
        char_offset += len(line) + 1

    for line_num, line in enumerate(lines):
        if line_num in pos_set and camelot_idx < len(camelot_mds):
            # Insert camelot table here
            result_lines.append("")  # blank line before table
            result_lines.append(camelot_mds[camelot_idx])
            result_lines.append("")  # blank line after table
            camelot_idx += 1
        else:
            result_lines.append(line)

    # Append any remaining camelot tables at the end
    while camelot_idx < len(camelot_mds):
        result_lines.append("")
        result_lines.append(camelot_mds[camelot_idx])
        result_lines.append("")
        camelot_idx += 1

    return "\n".join(result_lines)


# ---------------------------------------------------------------------------
# PDF extraction (pymupdf4llm for text + camelot for tables)
# ---------------------------------------------------------------------------

def extract_pdf_content(raw_bytes: bytes) -> Tuple[str, list]:
    """Extract text and structured tables from PDF bytes as clean Markdown.

    Uses a two-pass approach:
    1. pymupdf4llm.to_markdown() for high-quality text extraction that
       preserves headings, bold/italic, and lists.
    2. camelot-py (lattice mode) for accurate table extraction that
       correctly handles merged cells in HKEx regulatory forms.

    Falls back gracefully: pymupdf4llm -> basic PyMuPDF for text,
    camelot -> pymupdf4llm inline tables for table extraction.
    """
    if not PYMUPDF_AVAILABLE:
        return "", []

    try:
        # --- Pass 1: Text extraction via pymupdf4llm ---
        doc = pymupdf.open(stream=raw_bytes, filetype="pdf")

        if PYMUPDF4LLM_AVAILABLE:
            md_text = pymupdf4llm.to_markdown(doc)
            doc.close()
            log(f"    pymupdf4llm extracted {len(md_text)} chars")
        else:
            pages_text: list = []
            for page in doc:
                page_text = page.get_text("text") or ""
                if page_text.strip():
                    pages_text.append(page_text)
            doc.close()
            md_text = "\n\n---\n\n".join(pages_text)
            log(f"    PyMuPDF fallback extracted {len(md_text)} chars")

        # Post-process: clean headers/footers, normalize formatting
        md_text = _clean_markdown(md_text)

        # --- Pass 2: Table extraction via camelot (preferred) ---
        tables_json = _extract_tables_with_camelot(raw_bytes)

        if tables_json:
            log(f"    Using {len(tables_json)} camelot tables "
                f"(accuracy: {', '.join(str(t['accuracy']) + '%' for t in tables_json)})")
            # Substitute pymupdf4llm inline tables with cleaner camelot tables
            md_text = _substitute_tables(md_text, tables_json)
            md_text = _clean_markdown(md_text)  # Re-clean after substitution
            log(f"    Substituted inline tables with camelot output")
        else:
            # Fallback: extract table metadata from pymupdf4llm Markdown
            tables_json = _extract_tables_from_md(md_text)
            if tables_json:
                log(f"    Camelot unavailable/empty, using {len(tables_json)} "
                    f"pymupdf4llm inline tables (may have merged-cell artifacts)")
            else:
                log(f"    No tables found in PDF")

        return md_text, tables_json

    except Exception as e:
        log(f"  PDF extraction error: {e}")
        return "", []


# ---------------------------------------------------------------------------
# HTML extraction (preserving heading structure)
# ---------------------------------------------------------------------------

def extract_html_content(raw_bytes: bytes) -> Tuple[str, list]:
    """Extract text and tables from HTML bytes as clean Markdown.

    Preserves heading hierarchy (h1-h6 -> # tags), converts HTML tables to
    Markdown tables, and maintains paragraph structure. Avoids duplicate
    content from nested elements by tracking already-processed nodes.
    """
    if not BS4_AVAILABLE:
        return "", []

    tables_json: list = []
    try:
        soup = BeautifulSoup(raw_bytes, "html.parser")

        # Remove non-content elements
        for s in soup(["script", "style", "nav", "footer", "header", "noscript"]):
            s.decompose()

        md_parts: list = []
        table_index = 0
        processed_tables: set = set()

        root = soup.body if soup.body else soup

        def _process_element(el) -> None:
            """Recursively process an element, skipping already-handled tables."""
            nonlocal table_index

            if not isinstance(el, Tag):
                return

            tag = el.name

            # Headings -> Markdown headings
            if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
                level = int(tag[1])
                heading_text = squash_ws(el.get_text())
                if heading_text:
                    md_parts.append(f"\n\n{'#' * level} {heading_text}\n")
                return

            # Paragraphs -> separated text blocks
            if tag == "p":
                p_text = squash_ws(el.get_text())
                if p_text:
                    md_parts.append(f"\n\n{p_text}\n")
                return

            # List items
            if tag == "li":
                li_text = squash_ws(el.get_text())
                if li_text:
                    parent = el.parent
                    if parent and parent.name == "ol":
                        siblings = list(parent.find_all("li", recursive=False))
                        idx = siblings.index(el) + 1 if el in siblings else 1
                        md_parts.append(f"\n{idx}. {li_text}")
                    else:
                        md_parts.append(f"\n- {li_text}")
                return

            # Tables -> Markdown tables (only process innermost data tables)
            if tag == "table":
                if id(el) in processed_tables:
                    return
                processed_tables.add(id(el))

                nested_tables = el.find_all("table")
                if nested_tables:
                    for child in el.children:
                        if isinstance(child, Tag):
                            _process_element(child)
                    return

                table_index += 1
                rows_data: list = []
                headers: list = []
                for tr in el.find_all("tr"):
                    cells = [squash_ws(td.get_text()) for td in tr.find_all(["td", "th"])]
                    if not headers:
                        headers = cells
                    else:
                        rows_data.append(cells)
                if headers and rows_data:
                    md = _table_to_markdown(headers, rows_data)
                    if md:
                        md_parts.append(f"\n\n{md}\n")
                        tables_json.append({
                            "tableIndex": table_index,
                            "headers": headers,
                            "rowCount": len(rows_data),
                            "markdown": md,
                        })
                elif headers:
                    text = " | ".join(c for c in headers if c.strip())
                    if text:
                        md_parts.append(f"\n\n{text}\n")
                return

            # Container elements -> recurse into children
            if tag in ("div", "span", "td", "th", "tr", "tbody", "thead",
                       "tfoot", "section", "article", "main", "aside",
                       "ul", "ol", "dl", "dd", "dt", "figure",
                       "figcaption", "blockquote", "form", "fieldset"):
                for child in el.children:
                    if isinstance(child, Tag):
                        _process_element(child)
                    elif hasattr(child, 'string') and child.string:
                        text = child.string.strip()
                        if text:
                            md_parts.append(f" {text}")
                return

            # Line breaks
            if tag == "br":
                md_parts.append("\n")
                return

        _process_element(root)
        raw_md = "".join(md_parts)

        # If structured extraction produced very little, fall back to plain text
        if len(raw_md.strip()) < 50:
            raw_md = soup.get_text(separator="\n\n")

        md_text = _clean_markdown(raw_md)
        log(f"    HTML extracted {len(md_text)} chars, {len(tables_json)} tables")
        return md_text, tables_json

    except Exception as e:
        log(f"  HTML extraction error: {e}")
        return "", []


# ---------------------------------------------------------------------------
# Excel extraction (structured Markdown with sheet headings and tables)
# ---------------------------------------------------------------------------

def extract_excel_content(raw_bytes: bytes) -> Tuple[str, list]:
    """Extract text and tables from Excel bytes as clean Markdown.

    Each sheet becomes a section with a heading, and data is formatted as
    a Markdown table. Key-value style sheets (2 columns) are formatted
    as definition lists for readability.
    """
    if not OPENPYXL_AVAILABLE:
        return "", []

    tables_json: list = []
    all_parts: list = []
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(raw_bytes), read_only=True, data_only=True
        )
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows_data: list = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(c.strip() for c in cells):
                    rows_data.append(cells)
            if not rows_data:
                continue

            # Sheet heading
            sheet_md = f"## Sheet: {sheet_name}\n\n"

            headers = rows_data[0]
            data_rows = rows_data[1:]

            # Detect key-value layout (2 columns, first column is labels)
            is_kv = (
                len(headers) == 2
                and len(data_rows) >= 2
                and all(row[0].strip() for row in data_rows if len(row) >= 2)
            )

            if is_kv:
                kv_lines: list = []
                label_h = squash_ws(headers[0]) if headers[0].strip() else "Field"
                value_h = squash_ws(headers[1]) if headers[1].strip() else "Value"
                kv_lines.append(f"**{label_h}** | **{value_h}**")
                kv_lines.append("---|---")
                for row in data_rows:
                    key = squash_ws(row[0]) if len(row) > 0 else ""
                    val = squash_ws(row[1]) if len(row) > 1 else ""
                    kv_lines.append(f"{key} | {val}")
                sheet_md += "\n".join(kv_lines)
                tables_json.append({
                    "tableIndex": sheet_idx,
                    "sheetName": sheet_name,
                    "headers": [label_h, value_h],
                    "rowCount": len(data_rows),
                    "markdown": "\n".join(kv_lines),
                })
            else:
                md = _table_to_markdown(headers, data_rows)
                if md:
                    sheet_md += md
                    tables_json.append({
                        "tableIndex": sheet_idx,
                        "sheetName": sheet_name,
                        "headers": [squash_ws(str(h)) for h in headers],
                        "rowCount": len(data_rows),
                        "markdown": md,
                    })

            all_parts.append(sheet_md)

        wb.close()
    except Exception as e:
        log(f"  Excel extraction error: {e}")

    md_text = _clean_markdown("\n\n".join(all_parts))
    log(f"    Excel extracted {len(md_text)} chars, {len(tables_json)} tables")
    return md_text, tables_json


# ---------------------------------------------------------------------------
# Router
# ---------------------------------------------------------------------------

def extract_content_with_tables(
    raw_bytes: bytes, doc_url: str
) -> Tuple[str, list]:
    """Route to the appropriate extractor based on file extension."""
    u = doc_url.lower().split("?")[0].split("#")[0]
    if u.endswith(".pdf"):
        return extract_pdf_content(raw_bytes)
    if u.endswith(".htm") or u.endswith(".html"):
        return extract_html_content(raw_bytes)
    if u.endswith(".xlsx") or u.endswith(".xls"):
        return extract_excel_content(raw_bytes)
    return "", []
